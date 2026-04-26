
import streamlit as st
from datetime import datetime
import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re

st.set_page_config(page_title="DAMO Bore Log Tool", layout="wide")

st.title("DAMO Bore Log Generator")

input_text = st.text_area("Paste Bore Input", height=300)


def parse_input(text):
    bores = []
    blocks = re.split(r'\n\s*\n', text.strip())
    for block in blocks:
        lines = [l.strip() for l in block.split("\n") if l.strip()]
        bore = {}
        m = re.match(r'(BR|PL)(\d+)\s*=\s*(\d+)', lines[0])
        if not m:
            continue
        bore["type"] = m.group(1)
        bore["name"] = f"{m.group(1)}{m.group(2)}"
        bore["footage"] = int(m.group(3))
        bore["lc"] = {}
        bore["eop_range"] = None  # FIX: default to None so PL without EOP won't crash
        bore["floats"] = []

        for line in lines[1:]:
            if line.startswith("LC"):
                parts = line.replace("LC", "").split("=")
                rods = [int(x.strip()) for x in parts[0].split(",")]
                names = [x.strip() for x in parts[1].split(",")]
                for r, n in zip(rods, names):
                    bore["lc"][r] = n

            elif line.startswith("EOP"):
                eop_match = re.search(r'(\d+)-(\d+)', line)
                if eop_match:
                    bore["eop_range"] = (int(eop_match.group(1)), int(eop_match.group(2)))

                        elif line.startswith("Depth"):
                nums = list(map(int, re.findall(r'\d+', line)))
                if bore["type"] == "BR":
                    bore["depth_range"] = (nums[0] * 12 + nums[1], nums[2] * 12 + nums[3])
                else:
                    if len(nums) >= 2:
                        bore["depth_flat"] = nums[0] * 12 + nums[1]
                    else:
                        bore["depth_flat"] = nums[0] * 12

            elif line.startswith("Float"):
                # parse: Float 1 @ rod 6-22 = 3'7"
                rod_match   = re.search(r'rod\s*(\d+)-(\d+)', line)
                depth_match = re.search(r'=\s*(\d+)\'(\d+)', line)
                if rod_match and depth_match:
                    rod_start  = int(rod_match.group(1))
                    rod_end    = int(rod_match.group(2))
                    depth_inch = int(depth_match.group(1)) * 12 + int(depth_match.group(2))
                    bore["floats"].append({
                        "rod_start": rod_start,
                        "rod_end":   rod_end,
                        "depth":     depth_inch
                    })


        bores.append(bore)
    return bores


def inches_to_ft_in(val):
    return val // 12, val % 12


def rods_from_footage(ft):
    return ft // 10 + (1 if ft % 10 else 0)


def generate_eop(rods, lo, hi):
    eop = {}
    val = random.randint(lo, hi)
    for r in range(5, rods + 1, 5):
        val = max(lo, min(hi, val + random.choice([-1, 0, 1])))
        eop[r] = val
    return eop


def generate_br_depths(rods, low, high, lc, floats):
    depths = []
    current = random.randint(low, low + 3)

    # ── STEP A: build float schedule ──────────────────────────────
    # if no floats provided, randomly generate 1-3 float zones
    if not floats:
        schedule = []
        rod = random.randint(4, 8)  # first float starts early
        for _ in range(random.randint(1, 3)):
            if rod >= rods:
                break
            duration   = random.randint(10, 20)
            rod_end    = min(rod + duration, rods)
            float_depth = random.randint(low + 3, high - 3)
            schedule.append({
                "rod_start": rod,
                "rod_end":   rod_end,
                "depth":     float_depth
            })
            rod = rod_end + random.randint(3, 8)  # gap between floats
    else:
        # use provided floats, clamp depth to valid range
        schedule = []
        for f in floats:
            clamped = max(low, min(high, f["depth"]))
            schedule.append({
                "rod_start": f["rod_start"],
                "rod_end":   f["rod_end"],
                "depth":     clamped
            })

    # ── STEP B: generate depths rod by rod ───────────────────────
    float_target = None  # current float we are heading toward or in

    for rod_index in range(rods):
        rod_num = rod_index + 1  # rod numbers start at 1

        # check if this rod is inside a float zone
        in_float = None
        for f in schedule:
            if f["rod_start"] <= rod_num <= f["rod_end"]:
                in_float = f
                break

        if in_float:
            float_target = in_float["depth"]

            # gradually ramp toward float target if not there yet
            if abs(current - float_target) > 3:
                step = random.randint(1, 4)
                if float_target > current:
                    current += step
                else:
                    current -= step
                current = max(low, min(high, current))

            else:
                # inside float zone — drift ±1-3 inches around target
                drift     = random.randint(-3, 3)
                candidate = max(low, min(high, float_target + drift))

                # probability decay on repeats
                repeat_count = 0
                if len(depths) >= 1 and depths[-1] == candidate:
                    repeat_count = 1
                if len(depths) >= 2 and depths[-1] == depths[-2] == candidate:
                    repeat_count = 2
                if len(depths) >= 3 and depths[-1] == depths[-2] == depths[-3] == candidate:
                    # force a change
                    while candidate == current:
                        candidate = max(low, min(high, current + random.choice([-3,-2,-1,1,2,3])))
                else:
                    repeat_chance = 0.70 ** repeat_count
                    if random.random() > repeat_chance:
                        candidate = max(low, min(high, current + random.choice([-3,-2,-1,1,2,3])))

                current = candidate

        else:
            # outside float zone — check if next float is coming up
            next_float = None
            for f in schedule:
                if f["rod_start"] > rod_num:
                    next_float = f
                    break

            if next_float:
                # ramp gradually toward next float target
                target = next_float["depth"]
                step   = random.randint(1, 4)
                if abs(current - target) > 3:
                    if target > current:
                        current += step
                    else:
                        current -= step
                else:
                    # close enough — small random drift
                    current += random.choice([-2, -1, 0, 1, 2])
                current = max(low, min(high, current))

            else:
                # past all floats — free random drift
                current += random.choice([-3,-2,-1,0,1,2,3])
                current  = max(low, min(high, current))

        depths.append(current)

    # ── STEP C: LC ramp ───────────────────────────────────────────
    # for each LC rod, ramp gradually toward surface depth over 5 rods
    for r in lc:
        if r - 1 >= len(depths):
            continue

        lc_target   = random.randint(36, 42)   # target depth at LC rod (3'0"–3'6")
        ramp_start  = max(0, r - 6)            # start ramping 5 rods before LC rod
        ramp_end    = r - 1                    # LC rod index (0-based)

        # figure out what depth we're coming from
        start_depth = depths[ramp_start]
        total_steps = ramp_end - ramp_start    # how many rods to cover

        for i, idx in enumerate(range(ramp_start, ramp_end + 1)):
            if idx >= len(depths):
                break

            # linear interpolation — smoothly blend from start_depth to lc_target
            progress       = i / total_steps
            blended        = start_depth + (lc_target - start_depth) * progress
            nudge          = random.randint(-1, 1)    # small natural variation
            depths[idx]    = max(36, min(high, int(blended) + nudge))

        # make sure the LC rod itself lands in range
        depths[ramp_end] = lc_target

    return depths[:rods]



def generate_pl_depths(rods, flat):
    return[flat] * rods
    

def validate_depths(depths, bore_name, lc):
    violations = []
    last_triple_end = -999

    # ── CHECK 1: repeat rules ─────────────────────────────────────
    for i in range(2, len(depths)):
        if depths[i] == depths[i-1] == depths[i-2]:
            triple_start = i - 2

            if triple_start - last_triple_end < 6:
                violations.append(
                    f"  Rod {triple_start+1}: triple repeat too soon "
                    f"(only {triple_start - last_triple_end} lines since last)"
                )
            last_triple_end = i

            if i + 1 < len(depths) and depths[i+1] == depths[i]:
                violations.append(
                    f"  Rod {triple_start+1}: 4-in-a-row found ({depths[i]}\")"
                )

    # ── CHECK 2: LC rod depth ─────────────────────────────────────
    for r in lc:
        if r - 1 >= len(depths):
            continue
        lc_depth = depths[r - 1]
        if not (36 <= lc_depth <= 42):
            ft  = lc_depth // 12
            inc = lc_depth % 12
            violations.append(
                f"  Rod {r} (LC): depth {ft}'{inc}\" is outside 3'0\"–3'6\" target"
            )

    # ── CHECK 3: LC ramp approach ─────────────────────────────────
    for r in lc:
        ramp_start = max(0, r - 6)
        ramp_end   = r - 1

        if ramp_end >= len(depths) or ramp_start >= len(depths):
            continue

        start_depth = depths[ramp_start]
        end_depth   = depths[ramp_end]

        # if it was ramping down, start should be higher than end
        # allow a small tolerance of 3 inches in case it was already near surface
        if start_depth > 42 and end_depth > start_depth:
            violations.append(
                f"  Rod {r} (LC): ramp did not descend — "
                f"started at {start_depth}\" ended at {end_depth}\""
            )

        # check no single step in the ramp jumped more than 6 inches
        for i in range(ramp_start + 1, ramp_end + 1):
            if i >= len(depths):
                break
            jump = abs(depths[i] - depths[i - 1])
            if jump > 6:
                violations.append(
                    f"  Rod {r} (LC): ramp jump too large at rod {i+1} "
                    f"({jump}\" in one step)"
                )

    if violations:
        return f"{bore_name}: ❌ FAILED\n" + "\n".join(violations)
    else:
        return f"{bore_name}: ✅ all rules pass"


def build_excel(bores):
    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    title_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    number_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    thin = Side(style='thin')
    box_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    validation_results = []

    for bore in bores:
        rods = rods_from_footage(bore["footage"])

        if bore["type"] == "BR":
            depths = generate_br_depths(rods, *bore["depth_range"], bore["lc"], bore["floats"])
            result = validate_depths(depths, bore["name"])
            st.write(result)
        else:
            depths = generate_pl_depths(rods, bore["depth_flat"])

        # FIX: only generate EOP if range was provided
        if bore["eop_range"]:
            eop = generate_eop(rods, *bore["eop_range"])
        else:
            eop = {}

        validation_result = validate_depths(depths, bore["name"], bore["lc"])
        validation_results.append(validation_result)


        for page_start in range(0, rods, 74):
            ws = wb.create_sheet(f"{bore['name']}_{page_start // 74 + 1}")

            # FIX: clean column dimension references (removed invisible unicode chars)
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 28
            ws.column_dimensions['C'].width = 6
            ws.column_dimensions['D'].width = 8
            ws.column_dimensions['E'].width = 8
            ws.column_dimensions['F'].width = 6
            ws.column_dimensions['G'].width = 28
            ws.column_dimensions['H'].width = 6
            ws.column_dimensions['I'].width = 8
            ws.column_dimensions['J'].width = 8

            ws.merge_cells("A1:J1")
            ws["A1"] = "DAMO Bore Log"
            ws["A1"].font = Font(bold=True, size=14)
            ws["A1"].fill = title_fill
            ws["A1"].alignment = Alignment(horizontal="center")

            headers = ["#", "Location Description", "EOP", "Depth (ft)", "Depth (in)"]

            for col in range(5):
                # LEFT SIDE (A–E)
                cell_left = ws.cell(row=3, column=col + 1, value=headers[col])
                cell_left.fill = number_fill if col == 0 else header_fill
                cell_left.font = Font(bold=True)
                cell_left.alignment = Alignment(horizontal="center")
                cell_left.border = box_border

                # RIGHT SIDE (F–J)
                cell_right = ws.cell(row=3, column=col + 6, value=headers[col])
                cell_right.fill = number_fill if col == 0 else header_fill
                cell_right.font = Font(bold=True)
                cell_right.alignment = Alignment(horizontal="center")
                cell_right.border = box_border

            for i in range(37):          # <- loop starts here
                left = page_start + i
                right = page_start + i + 37
                row_excel = 4 + i

                # LEFT SIDE
                for col in range(1, 11):
                    ws.cell(row=row_excel, column=col).border = box_border
                    
                if left < rods:
                    rod = left + 1
                    ft, inch = inches_to_ft_in(depths[left])
                    c = ws.cell(row=row_excel, column=1, value=rod)
                    c.fill = number_fill
                    c.border = box_border
                    ws.cell(row=row_excel, column=2, value=bore["lc"].get(rod, ""))
                    ws.cell(row=row_excel, column=3, value=eop.get(rod, ""))
                    ws.cell(row=row_excel, column=4, value=ft)
                    ws.cell(row=row_excel, column=5, value=inch)

                # RIGHT SIDE
                if right < rods:
                    rod = right + 1
                    ft, inch = inches_to_ft_in(depths[right])
                    c = ws.cell(row=row_excel, column=6, value=rod)
                    c.fill = number_fill
                    c.border = box_border
                    ws.cell(row=row_excel, column=7, value=bore["lc"].get(rod, ""))
                    ws.cell(row=row_excel, column=8, value=eop.get(rod, ""))
                    ws.cell(row=row_excel, column=9, value=ft)
                    ws.cell(row=row_excel, column=10, value=inch)

            # Total line on last page
            if page_start + 74 >= rods:
                ws.cell(row=42, column=9, value=f"Total: {bore['footage']}'")

    filename = f"DAMO_OUTPUT_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    return filename


if st.button("Generate Bore Log"):
    if input_text.strip():
        try:
            bores = parse_input(input_text)
            if not bores:
                st.error("No valid bores found. Check your input format.")
            else:
                file = build_excel(bores)
                with open(file, "rb") as f:
                    st.download_button("Download Excel", f, file_name=file)
                st.success(f"Generated {len(bores)} bore(s) successfully!")
                st.subheader("Validation Results")
                for result in validation_results:
                    if "FAILED" in result:
                        st.error(result)
                    else:
                        st.success(result)

        except Exception as e:
            st.error(f"Error generating bore log: {e}")
    else:
        st.warning("Paste input first.")
